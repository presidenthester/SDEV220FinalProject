import tkinter as Tk
from tkinter import *
from tkinter import ttk
import os
import openpyxl


# This list is used in the combobox below.  It is a complete list of every NFL team.
teams = [
    "Arizona Cardinals",
    "Atlanta Falcons",
    "Baltimore Ravens",
    "Buffalo Bills",
    "Carolina Panthers",
    "Chicago Bears",
    "Cincinnati Bengals",
    "Cleveland Browns",
    "Dallas Cowboys",
    "Denver Broncos",
    "Detroit Lions",
    "Greenbay Packers",
    "Houston Texans",
    "Indianapolis Colts",
    "Jacksonville Jaguars",
    "Kansas City Chiefs",
    "Las Vegas Raiders",
    "Los Angeles Chargers",
    "Los Angelous Rams",
    "Miami Dolphins",
    "Minnesota Vikings",
    "New England Patriots",
    "New Orleans Saints",
    "New York Giants",
    "New York Jets",
    "Philadelphia Eagles",
    "Pittsburgh Steelers",
    "San Francisco 49ers",
    "Seattle Seahawks",
    "Tampa Bay Buccaneers",
    "Tennessee Titans",
    "Washington Commanders"
    
]


# This class contains functions that set up all of the windows in the program
class Panels:

    def __init__(self):
        pass

    def MainWindow(self):
        # Define main panel
        main = Tk()
        main.geometry("1920x1080")
        main.title("Cap Rankngs")
        main.resizable(True, True)
        main.iconbitmap("C:\\Users\\ericj_dmnhny4\\Desktop\\Cap_ranking_Project\\Images\\vince-lombardi-football-trophy.ico")


        # Define Canvas
        main_canvas = Canvas(main, width=1920, height=1080)
        main_canvas.pack(fill="both", expand=True)
        main_canvas.configure(bg='#516f75')
        main_canvas.pack()

        # get the screen dimensions
        screen_width = main.winfo_screenwidth()
        screen_height = main.winfo_screenheight()

        # calculate the position of the window
        window_width = 1920
        window_height = 1080
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2

        # set the geometry of the window
        main.geometry(f"{window_width}x{window_height}+{x}+{y}")

        # Create custom TTK styles
        style = ttk.Style()
        style.configure('TButton', borderwidth=0, relief="groove", background="black", foreground="#2c3030", padding=10, width=50, height=50)
        style.map('TButton', background=[('active', 'black')], foreground=[('active', '#06313b')])

        # Define text
        message_1 = Message(main, text="Salary cap for 2023")
        message_1.config(bg='#516f75', fg='#3bb9f7', font=('times', 24, 'bold'), width=600)
        message_2 = Message(main, text="$224.8 million")
        message_2.config(bg='#516f75', fg='#3bb9f7', font=('times', 36, 'bold'), width=600)

        # Define buttons
        player_entry_button = ttk.Button(main, text= 'Enter Player Profile',style='TButton', command= pan.player_profile_entry)
        view_players = ttk.Button(main, text= 'View Players', style='TButton', command=pan.view_players_panel)

        # Place buttons on canvas
        main_canvas.create_window(680, 540, window=player_entry_button)
        main_canvas.create_window(1240, 540, window=view_players)
        main_canvas.create_window(960, 100, window=message_1)
        main_canvas.create_window(960, 150, window=message_2)

        mainloop()

        
    def player_profile_entry(self):

        player_profile_entry = Toplevel()
        player_profile_entry.title("Player Profile Entry")
        player_profile_entry.iconbitmap("C:\\Users\\ericj_dmnhny4\\Desktop\\Cap_ranking_Project\\Images\\vince-lombardi-football-trophy.ico")

        # Top level frames and widgets definition
        ppe_frame = Frame(player_profile_entry, width=750, height=500)
        ppe_frame.pack(fill= "both", expand= True)

        player_info_frame = LabelFrame(ppe_frame,text= "Player Information")
        player_info_frame.grid( row= 0, column= 0, padx= 20, pady= 20 )

        
        # Frame widget definition and placement
        fname_label = Label(player_info_frame, text= "First Name")
        fname_label.grid(row= 0, column= 0)

        lname_label = Label(player_info_frame, text="Last Name")
        lname_label.grid(row= 0 , column= 1)

        self.fname_entry = Entry(player_info_frame)
        self.fname_entry.grid(row= 1, column= 0)

        self.lname_entry = Entry(player_info_frame)
        self.lname_entry.grid(row= 1, column= 1)

        team_label = Label(player_info_frame, text="Current Team")
        self.team_combobox = ttk.Combobox(player_info_frame, values= teams)
        team_label.grid(row=0, column=2)
        self.team_combobox.grid(row=1, column=2)

        age_label = Label(player_info_frame, text="Age")
        self.age_spinbox = Spinbox(player_info_frame, from_=18, to= 50)
        age_label.grid(row= 2, column= 0)
        self.age_spinbox.grid(row=3, column=0)

        current_salary_label = Label(player_info_frame, text="Current Salary")
        self.current_salary_entry = Entry(player_info_frame)
        current_salary_label.grid(row=2, column=1)
        self.current_salary_entry.grid(row=3, column=1)

        projected_salary_label = Label(player_info_frame, text="Projected Salary")
        self.projected_salary_entry = Entry(player_info_frame)
        projected_salary_label.grid(row=2, column=2)
        self.projected_salary_entry.grid(row=3, column=2)

        # Loop used to make a nicer looking form setting padding for X and Y padding
        for widget in player_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)
        
        # 2nd Frame for injury history and widgets seperating them
        injury_info_frame = LabelFrame(ppe_frame,text= "Injury Information")
        injury_info_frame.grid( row= 1, column= 0, sticky="news", padx=20, pady=20)

        num_injuries_label = Label( injury_info_frame, text="Number of Injuries")
        num_injuries_label.grid(row=0, column=0)

        missed_games_label = Label(injury_info_frame, text="Games Missed")
        missed_games_label.grid(row=0, column= 1)

        self.num_injuries_entry = Entry(injury_info_frame)
        self.num_injuries_entry.grid(row=1, column=0)

        self.missed_games_entry = Entry(injury_info_frame)
        self.missed_games_entry.grid(row=1, column=1)

        for widget in injury_info_frame.winfo_children():
            widget.grid_configure(padx=10, pady=5)

        

        enter_player_button = Button(ppe_frame, text="Enter Player", command=lambda: [de.get_player_info(), ds.create_sheet()])
        enter_player_button.grid(row=2, column=0, padx=20, pady=20)


        
    def view_players_panel(self):
        players_viewer = Toplevel()
        players_viewer.title("Player Profile Entry")
        players_viewer.iconbitmap("C:\\Users\\ericj_dmnhny4\\Desktop\\Cap_ranking_Project\\Images\\vince-lombardi-football-trophy.ico")

        vpp_frame = Frame(players_viewer, width=1920, height=1080)
        vpp_frame.pack(fill= "both", expand= True)

        self.path = "C:\\Users\\ericj_dmnhny4\\Desktop\\Cap_ranking_Project\\players.xlsx"

        if not os.path.exists(self.path):
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.heading = ["First Name", "Last Name", "Age", "Team", "Current Salary", 
                       "Projected Salary", "No. Injuries", "Games Missed", "Score"]
            
            self.sheet.append(self.heading)
            self.workbook.save(self.path)

        self.workbook = openpyxl.load_workbook(self.path)
        self.sheet = self.workbook.active

        player_values   = list(self.sheet.values)
        cols            = player_values[0]
        plant           = ttk.Treeview(vpp_frame, columns= cols, show="headings")

        for column_name in cols:
            plant.heading(column_name, text=column_name)
        plant.pack(expand=True, fill='y')

        for tup_val in player_values[1:]:
            plant.insert('', ttk.tkinter.END, values=tup_val)

        
# This class gets player information and creates a database from it
class DataEntry:
    def __init__(self):
        pass
    
    # This funnctions retrieves the data entered in player info form
    def get_player_info(self):

        self.firstname  =   pan.fname_entry.get()
        self.lastname   =   pan.lname_entry.get()
        self.team       =   pan.team_combobox.get()
        self.age        =   pan.age_spinbox.get()
        self.cur_sal    =   pan.current_salary_entry.get()
        self.proj_sal   =   pan.projected_salary_entry.get()
        self.injuries   =   pan.num_injuries_entry.get()
        self.missed     =   pan.missed_games_entry.get()

        
        self.age_base           =   self.age
        self.cur_sal_base       =   self.cur_sal
        self.proj_sal_base      =   self.proj_sal
        self.injuries_base      =   self.injuries
        self.missed_games_base  =   self.missed
        self.salary_score       =   0
        self.injury_score       =   0
        self.missed_games_score =   0
        self.total_injury_score =   0
        self.player_score       =   0
        


        # Gets the current data entry for age and converts it to a score of 1 to 10 (A player between the age of 23 and 30 has the best age score fo 10)
        if de.age_base >= str(18) and de.age_base <= str(22):
            de.age_base = 7

        elif de.age_base >= str(23) and de.age_base <= str(30):
            de.age_base = 10

        elif de.age_base >= str(31) and de.age_base <= str(33):
            de.age_base = 8

        elif de.age_base >= str(34) and de.age_base <= str(36):
            de.age_base = 7

        elif de.age_base >= str(37) and de.age_base <= str(50):
            de.age_base = 5

        
             # Comapares a players current and projected salary and applies a score 
        if str(self.proj_sal_base) <= str(self.cur_sal_base):
            self.salary_score = 10

        elif str(self.proj_sal_base) <= str((float(self.cur_sal_base) * 1.05)):
            self.salary_score = 9

        elif str(self.proj_sal_base) <= str((float(self.cur_sal_base) * 1.10)):
            self.salary_score = 8

        elif str(self.proj_sal_base) <= str((float(self.cur_sal_base) * 1.15)):
            self.salary_score = 7

        elif str(self.proj_sal_base) <= str((float(self.cur_sal_base) * 1.20)):
            self.salary_score = 6

        else:
            self.salary_score = 5

           
        # Applies a score to the amount of injuries a player has had
        if self.injuries_base >= str(0) and self.injuries_base <= str(1):
            self.injury_score = 0
        
        elif self.injuries_base >= str(2) and self.injuries_base <= str(3):
            self.injury_score = 2
        
        elif self.injuries_base <= str(5) or self.injuries_base == str(4):
            self.injury_score = 4

        elif self.injuries_base > str(5):
            self.injury_score = 6

        # Applies a score to the number games a player has missed

        if self.missed_games_base >= str(0) and self.missed_games_base <= str(1):
            self.missed_games_score = 0

        elif self.missed_games_base >= str(2) and self.missed_games_base <= str(5):
            self.missed_games_score = 2

        elif self.missed_games_base <= str(10) or self.missed_games_base == str(6):
            self.missed_games_score = 4

        elif self.missed_games_base <= str(17) or self.missed_games_base == str(7):
            self.missed_games_score = 6

        elif self.missed_games_base >= str(18): 
            self.missed_games_score = 8

        

        self.player_score = (int(self.age_base) + int(self.salary_score)) - (self.missed_games_score) - (self.injury_score)

     
      
        
        print("First Name: ", self.firstname)
        print("Last Name: ", self.lastname)
        print("Age: ", self.age)
        print("Current Team: ", self.team)
        print("Current Salary: ", self.cur_sal)
        print("Projected Salary: ", self.proj_sal)
        print("Number of Injuries: ", self.injuries)
        print("Number of missed games: ", self.missed)
        print("Age Score: ", self.age_base)
        print("Salary Score: ", self.salary_score)
        print("Injury Score: ", self.injury_score)
        print("Games Missed Score: ", self.missed_games_score)
        print(self.total_injury_score)
        print(self.player_score)

# This class creates an excel file to enter and save entered data
class DataSheet:
    def __init__(self):
        self.sheet = []

    def create_sheet(self):
        self.filepath = "C:\\Users\\ericj_dmnhny4\\Desktop\\Cap_ranking_Project\\players.xlsx"

        if not os.path.exists(self.filepath):
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.heading = ["First Name", "Last Name", "Age", "Team", "Current Salary", 
                       "Projected Salary", "No. Injuries", "Games Missed", "Score"]
            
            self.sheet.append(self.heading)
            self.workbook.save(self.filepath)

        self.workbook = openpyxl.load_workbook(self.filepath)
        self.sheet = self.workbook.active
        self.sheet.append([de.firstname, de.lastname, de.age, de.team, de.cur_sal, 
                           de.proj_sal, de.injuries, de.missed, de.player_score])
        self.workbook.save(self.filepath)
        

pan =   Panels() 
de  =   DataEntry()
ds  =   DataSheet()
